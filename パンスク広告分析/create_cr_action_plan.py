#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# Color constants
DARK_BLUE = "1F3864"
LIGHT_BLUE = "BDD7EE"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FFE0E0"
LIGHT_GRAY = "F2F2F2"
DARK_GRAY = "D9D9D9"
DARK_ORANGE = "C55A11"
DARK_GREEN = "375623"
DARK_RED = "C00000"
WHITE = "FFFFFF"
MEDIUM_GRAY = "BFBFBF"

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_font(name="Arial", size=10, bold=False, italic=False, color="000000"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def make_border(style="thin"):
    side = Side(style=style, color="000000")
    return Border(left=side, right=side, top=side, bottom=side)

def make_align(horizontal="left", vertical="center", wrap=True):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def apply_header_style(cell, bg_color=LIGHT_BLUE, font_color="000000", bold=True, size=10, align="center"):
    cell.fill = make_fill(bg_color)
    cell.font = make_font(bold=bold, color=font_color, size=size)
    cell.alignment = make_align(horizontal=align)
    cell.border = make_border()

def apply_data_style(cell, bg_color=WHITE, bold=False, size=9, align="left", wrap=True):
    cell.fill = make_fill(bg_color)
    cell.font = make_font(bold=bold, size=size)
    cell.alignment = make_align(horizontal=align, wrap=wrap)
    cell.border = make_border()

def set_row_color(ws, row, col_start, col_end, color):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = make_fill(color)
        cell.border = make_border()

def apply_full_row_style(ws, row, col_start, col_end, bg_color, bold=False, size=9, align="left"):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = make_fill(bg_color)
        cell.font = make_font(bold=bold, size=size)
        cell.alignment = make_align(horizontal=align)
        cell.border = make_border()


# ==================== SHEET 1: 訴求軸サマリー ====================

def create_sheet1(wb):
    ws = wb.active
    ws.title = "訴求軸サマリー"

    # Row 1: Title
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = "パンスク Meta広告 CR訴求分析 アクションプラン"
    cell.fill = make_fill(DARK_BLUE)
    cell.font = make_font(bold=True, color=WHITE, size=14)
    cell.alignment = make_align(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Row 2: Subtitle
    ws.merge_cells("A2:I2")
    cell = ws["A2"]
    cell.value = "データ期間：2026/01/01〜2026/03/15 | 対象キャンペーン：ASC / ASC検証 / INT"
    cell.fill = make_fill(LIGHT_GRAY)
    cell.font = make_font(italic=True, size=10)
    cell.alignment = make_align(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Row 4: Section header
    ws.merge_cells("A4:I4")
    cell = ws["A4"]
    cell.value = "■ 全体パフォーマンスサマリー"
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[4].height = 18

    # Row 5: Campaign headers
    headers5 = ["キャンペーン", "CR数", "消化金額", "CV数", "CPA", "CTR", "評価", "", ""]
    for i, h in enumerate(headers5, 1):
        cell = ws.cell(row=5, column=i, value=h)
        apply_header_style(cell, bg_color=LIGHT_BLUE)

    # Campaign data rows 6-9
    campaign_data = [
        ("ASC キャンペーン", 54, "¥369,491", 59, "¥6,263", "3.04%", "✅ メイン軸"),
        ("ASC キャンペーン_検証", 9, "¥52,629", 7, "¥7,518", "1.79%", "🔬 テスト枠"),
        ("INTキャンペーン", 34, "¥25,740", 1, "¥25,740", "3.26%", "❌ 抜本見直し必要"),
        ("合計", 97, "¥447,860", 67, "¥6,684", "2.80%", "（全体）"),
    ]
    for i, (row_data, row_num) in enumerate(zip(campaign_data, range(6, 10))):
        is_total = row_num == 9
        bg = LIGHT_GRAY if is_total else WHITE
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            apply_data_style(cell, bg_color=bg, bold=is_total, size=9 if not is_total else 10)

    ws.row_dimensions[8].height = 18  # INT row

    # Row 11: Section header
    ws.merge_cells("A11:I11")
    cell = ws["A11"]
    cell.value = "■ フォーマット別パフォーマンス比較"
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[11].height = 18

    # Row 12: Format headers
    headers12 = ["フォーマット", "CR数", "CV数", "CPA", "CTR", "CVR", "評価", "示唆", ""]
    for i, h in enumerate(headers12, 1):
        cell = ws.cell(row=12, column=i, value=h)
        apply_header_style(cell, bg_color=LIGHT_BLUE)

    # Format data rows 13-15
    format_data = [
        ("カルーセル", 2, 2, "¥2,456", "4.06%", "3.03%", "✅ 最高効率", "本数を増やして検証すべき"),
        ("動画", 8, 24, "¥5,537", "1.84%", "1.79%", "✅ 最多CV", "CVRが最も安定。主力フォーマット"),
        ("画像（静止画）", 87, 41, "¥7,562", "3.32%", "0.90%", "⚠️ ばらつき大", "一部高効率CRあり。精査して絞り込む"),
    ]
    format_colors = [LIGHT_GREEN, LIGHT_GREEN, LIGHT_YELLOW]
    for row_data, row_num, color in zip(format_data, range(13, 16), format_colors):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            apply_data_style(cell, bg_color=color, size=9)

    # Row 17: Section header
    ws.merge_cells("A17:I17")
    cell = ws["A17"]
    cell.value = "■ 訴求3軸フレームワーク"
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[17].height = 18

    # Row 18: Framework headers
    headers18 = ["訴求軸", "コンセプト", "代表CR", "結果", "CTR傾向", "CVR傾向", "継続率への影響", "推奨方針", ""]
    for i, h in enumerate(headers18, 1):
        cell = ws.cell(row=18, column=i, value=h)
        apply_header_style(cell, bg_color=LIGHT_BLUE)

    # Framework data rows 19-21
    framework_data = [
        ("訴求軸A 生活提案型", "「毎朝の食卓が変わる」便利さ・豊かさ体験", "動画_08、動画_09、カルーセル系", "CV 24件 CPA¥5,209", "低め(1.8%)", "高め(1.79%)", "◎ 高継続率が期待できる", "✅ 最優先で強化"),
        ("訴求軸B 商品・品質型", "「職人のパン」品質・見た目・こだわり", "no85-4、no88-1、no91系、no62", "CV 28件以上 CPA¥2,358〜¥5,771", "高め(3〜5%)", "中〜高(1〜4%)", "○ 価値を理解した入会", "✅ 継続強化＋バリエーション追加"),
        ("訴求軸C 価格・お得感型", "「お得感」数量・価格を前面に", "no54系、no52系、no86、no87", "CV≒0 CTRは高い", "高め(4〜7%)", "ほぼ0%", "△ 解約理由1位が「価格高い」と相関", "❌ 抑制検討。LP改善前提で再テスト"),
    ]
    framework_colors = [LIGHT_GREEN, LIGHT_GREEN, "FFE0E0"]
    for row_data, row_num, color in zip(framework_data, range(19, 22), framework_colors):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            apply_data_style(cell, bg_color=color, size=9)
        ws.row_dimensions[row_num].height = 32

    # Row 23: Section header
    ws.merge_cells("A23:I23")
    cell = ws["A23"]
    cell.value = "■ 重要インサイト（バリアント分析）"
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[23].height = 18

    # Row 24: Variant headers
    headers24 = ["バリアント比較", "勝ちCR", "負けCR", "差異の示唆", "", "", "", "", ""]
    for i, h in enumerate(headers24, 1):
        cell = ws.cell(row=24, column=i, value=h)
        apply_header_style(cell, bg_color=LIGHT_BLUE)

    # Variant data rows 25-30
    variant_data = [
        ("no85シリーズ", "no85-4（23CV CPA¥5,771）", "no85-1/no85-3（CV0）", "同じベースで訴求/デザインの微差が結果を大きく左右。-4の要素を次期CRに踏襲"),
        ("no88シリーズ", "no88-1（2CV CPA¥2,358）", "no88-2（CV0）", "-1の要素（コピー/構成）が明確に機能。-2は停止"),
        ("no90シリーズ", "no90-2（1CV CPA¥1,015）", "no90-1（CV0 ×2キャンペーン）", "-2の差別点を分析して次期CRへ"),
        ("no91シリーズ", "no91-2（2CV 計）", "no91-1（1CV）", "両方機能するが-2がやや優勢"),
        ("no54シリーズ", "なし", "no54_a,b（高CTR・CVR≈0）", "CTRは6.9%と高いがCV0。LP到達後の離脱→LP改善が先決。訴求軸CのLP整合性問題"),
        ("no89シリーズ", "なし", "no89-1（3配信 全CV0）", "3つのキャンペーン全てでCV0。訴求自体が合っていない可能性。即停止検討"),
    ]
    for i, (row_data, row_num) in enumerate(zip(variant_data, range(25, 31))):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            if j <= 4:
                apply_data_style(cell, bg_color=bg, size=9)
        # Fill remaining cols
        for j in range(5, 10):
            cell = ws.cell(row=row_num, column=j)
            cell.fill = make_fill(bg)
            cell.border = make_border()
        ws.row_dimensions[row_num].height = 30

    # Column widths
    col_widths = {
        'A': 25, 'B': 28, 'C': 20, 'D': 20, 'E': 12, 'F': 12, 'G': 18, 'H': 40, 'I': 8
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    return ws


# ==================== SHEET 2: CR評価・アクション表 ====================

def create_sheet2(wb):
    ws = wb.create_sheet("CR評価・アクション表")

    # Title row
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "CR別評価 & アクション一覧"
    cell.fill = make_fill(DARK_BLUE)
    cell.font = make_font(bold=True, color=WHITE, size=14)
    cell.alignment = make_align(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Subtitle row
    ws.merge_cells("A2:L2")
    cell = ws["A2"]
    cell.value = "※ 評価基準：優先停止＝CV0かつ消化¥1,500以上 / 強化＝CV獲得かつCPA¥7,000以下"
    cell.fill = make_fill(LIGHT_GRAY)
    cell.font = make_font(italic=True, size=10)
    cell.alignment = make_align(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Headers row 3
    headers = ["広告名", "キャンペーン", "フォーマット", "訴求軸", "CV数", "CPA", "CTR", "CVR", "消化金額", "アクション", "優先度", "備考"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        apply_header_style(cell, bg_color=DARK_BLUE, font_color=WHITE)

    # CV獲得 CRs data
    cv_data = [
        ("動画_08", "ASC", "動画", "軸A 生活提案型", 23, "¥5,210", "1.85%", "1.79%", "¥119,819", "継続強化", "最高", "最多CV。動画フォーマットの主力"),
        ("no85-4", "ASC", "画像", "軸B 商品・品質型", 23, "¥5,771", "4.72%", "1.07%", "¥132,742", "継続+LP改善", "最高", "最多CV。CTR高いがCVR低め→LP改善でCPAが大幅改善の余地"),
        ("no88-1", "ASC", "画像", "軸B 商品・品質型", 2, "¥2,358", "3.22%", "3.92%", "¥4,715", "予算拡大", "高", "高効率。予算増額テスト推奨"),
        ("no62", "ASC", "画像", "軸B 商品・品質型", 1, "¥595", "7.09%", "10.0%", "¥595", "大幅拡大", "最高", "CPA¥595と最高効率。IMP141と極小→拡大優先"),
        ("no90-2", "ASC", "画像", "軸B 商品・品質型", 1, "¥1,015", "3.56%", "6.67%", "¥1,015", "予算拡大", "高", "高効率。同系列-1は全失敗→-2の要素を踏襲した新CR開発"),
        ("no44", "INT", "画像", "軸B 商品・品質型", 1, "¥1,436", "2.85%", "7.14%", "¥1,436", "ASCへ移行テスト", "高", "INT唯一のCV。ASCで試す価値あり"),
        ("no91-2", "ASC検証", "画像", "軸B 商品・品質型", 1, "¥1,442", "2.15%", "5.0%", "¥1,442", "継続", "高", "効率良好。ASCメインへの予算移行を検討"),
        ("2405_カルーセル02", "ASC", "カルーセル", "軸A 生活提案型", 1, "¥2,035", "3.86%", "3.03%", "¥2,035", "継続+新カルーセル開発", "高", "カルーセル形式の高効率を確認。新作開発推奨"),
        ("no98", "ASC検証", "画像", "軸B 商品・品質型", 1, "¥2,559", "1.33%", "2.44%", "¥2,559", "継続観察", "中", "効率良好だがCTR低め。継続して様子見"),
        ("no91-1", "ASC", "画像", "軸B 商品・品質型", 1, "¥2,598", "3.47%", "4.0%", "¥2,598", "継続", "中", ""),
        ("2405_カルーセル01", "ASC", "カルーセル", "軸A 生活提案型", 1, "¥2,878", "4.22%", "2.22%", "¥2,878", "継続", "中", "カルーセル2本とも成果。フォーマット有効性を確認"),
        ("no91-2", "ASC", "画像", "軸B 商品・品質型", 1, "¥4,039", "2.84%", "3.03%", "¥4,039", "継続", "中", ""),
        ("no56", "ASC検証", "画像", "軸B 商品・品質型", 2, "¥4,737", "1.36%", "1.41%", "¥9,474", "継続観察", "中", ""),
        ("動画_09", "ASC", "動画", "軸A 生活提案型", 1, "¥9,241", "1.72%", "1.43%", "¥9,241", "観察継続", "低", "CPA高め。動画_08への集中が先決"),
        ("no9", "ASC", "画像", "軸C 価格・お得感型", 3, "¥12,652", "2.63%", "0.58%", "¥37,957", "予算削減", "-", "CPA高い。消化が多い割にCVR低すぎ"),
        ("no48", "ASC検証", "画像", "軸C 価格・お得感型", 2, "¥13,134", "2.35%", "0.49%", "¥26,267", "停止検討", "-", "CPA高い"),
        ("no54_b", "ASC", "画像", "軸C 価格・お得感型", 1, "¥32,943", "4.50%", "0.20%", "¥32,943", "即停止", "最高(停止)", "CTR高いがCVR最悪。LP整合性なし"),
    ]

    # CV0 CRs data
    cv0_data = [
        ("no89-1", "ASC", "画像", "-", 0, "-", "1.23%", "0%", "¥1,673", "即停止", "-", "3キャンペーン全てCV0"),
        ("no6", "ASC", "画像", "-", 0, "-", "2.65%", "0%", "¥2,522", "停止", "-", ""),
        ("no88-2", "ASC", "画像", "-", 0, "-", "3.21%", "0%", "¥3,555", "停止", "-", "no88-1は機能→-2は不要"),
        ("no47", "INT", "画像", "-", 0, "-", "3.55%", "0%", "¥2,278", "停止", "-", ""),
        ("no86", "INT", "画像", "-", 0, "-", "3.57%", "0%", "¥2,342", "停止", "-", ""),
        ("no95", "ASC検証", "画像", "-", 0, "-", "1.23%", "0%", "¥2,889", "停止", "-", ""),
        ("no85-1", "ASC", "画像", "-", 0, "-", "3.69%", "0%", "¥2,061", "停止", "-", "no85-4が機能→-1は不要"),
        ("no87", "INT", "画像", "-", 0, "-", "4.29%", "0%", "¥1,530", "停止", "-", ""),
        ("no90-1", "ASC検証", "画像", "-", 0, "-", "2.41%", "0%", "¥1,319", "停止", "-", "no90-2が機能→-1は不要"),
        ("no54_b", "INT", "画像", "軸C", 0, "-", "4.09%", "0%", "¥3,763", "即停止", "-", "CTR高いがCV0。INTでも同様"),
    ]

    row_num = 4
    for row_data in cv_data:
        action = row_data[9]
        if action in ("継続強化", "予算拡大", "大幅拡大", "継続+LP改善", "継続+新カルーセル開発", "ASCへ移行テスト"):
            bg = LIGHT_GREEN
        elif action == "即停止":
            bg = LIGHT_RED
        elif action in ("停止", "停止検討", "予算削減"):
            bg = LIGHT_YELLOW
        elif row_num % 2 == 0:
            bg = WHITE
        else:
            bg = LIGHT_GRAY

        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        ws.row_dimensions[row_num].height = 28
        row_num += 1

    # Separator for CV0 section
    ws.merge_cells(f"A{row_num}:L{row_num}")
    cell = ws.cell(row=row_num, column=1, value="▼ CV0 停止対象CR")
    cell.fill = make_fill("FF9999")
    cell.font = make_font(bold=True, size=10, color=WHITE)
    cell.alignment = make_align(horizontal="center")
    row_num += 1

    for row_data in cv0_data:
        action = row_data[9]
        if action == "即停止":
            bg = LIGHT_RED
        elif action in ("停止", "停止検討"):
            bg = LIGHT_YELLOW
        else:
            bg = LIGHT_GRAY

        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        ws.row_dimensions[row_num].height = 24
        row_num += 1

    # Column widths
    col_widths = {
        1: 22, 2: 18, 3: 14, 4: 18, 5: 8, 6: 10, 7: 8, 8: 8, 9: 12, 10: 20, 11: 14, 12: 40
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"
    return ws


# ==================== SHEET 3: 訴求別深掘り分析 ====================

def create_sheet3(wb):
    ws = wb.create_sheet("訴求別深掘り分析")

    # Title row
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "訴求軸別 パフォーマンス深掘り"
    cell.fill = make_fill(DARK_BLUE)
    cell.font = make_font(bold=True, color=WHITE, size=14)
    cell.alignment = make_align(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "データ期間：2026/01/01〜2026/03/15"
    cell.fill = make_fill(LIGHT_GRAY)
    cell.font = make_font(italic=True, size=10)
    cell.alignment = make_align(horizontal="center")
    ws.row_dimensions[2].height = 18

    data_headers = ["CR名", "フォーマット", "CV数", "CPA", "CTR", "CVR", "消化金額", "評価"]

    row = 4

    # ---- Section A ----
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row=row, column=1, value="■ 訴求軸A：生活提案型（動画・カルーセル中心）")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    for i, h in enumerate(data_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_header_style(cell, bg_color=LIGHT_BLUE)
    row += 1

    axis_a_data = [
        ("動画_08", "動画", 23, "¥5,210", "1.85%", "1.79%", "¥119,819", "✅ 主力"),
        ("2405_カルーセル02", "カルーセル", 1, "¥2,035", "3.86%", "3.03%", "¥2,035", "✅ 高効率"),
        ("2405_カルーセル01", "カルーセル", 1, "¥2,878", "4.22%", "2.22%", "¥2,878", "✅"),
        ("動画_09", "動画", 1, "¥9,241", "1.72%", "1.43%", "¥9,241", "⚠️ 高CPA"),
    ]
    for i, row_data in enumerate(axis_a_data):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        row += 1

    # Subtotal A
    subtotal_a = ("軸A合計", "-", 26, "¥5,245（平均）", "2.29%", "1.69%", "¥134,973", "-")
    for j, val in enumerate(subtotal_a, 1):
        cell = ws.cell(row=row, column=j, value=val)
        apply_data_style(cell, bg_color=LIGHT_GRAY, bold=True, size=9)
    row += 2

    # ---- Section B ----
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row=row, column=1, value="■ 訴求軸B：商品・品質型（高効率画像CR）")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    for i, h in enumerate(data_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_header_style(cell, bg_color=LIGHT_BLUE)
    row += 1

    axis_b_data = [
        ("no85-4", "画像", 23, "¥5,771", "4.72%", "1.07%", "¥132,742", "✅ 最多CV"),
        ("no88-1", "画像", 2, "¥2,358", "3.22%", "3.92%", "¥4,715", "✅ 高効率"),
        ("no62", "画像", 1, "¥595", "7.09%", "10.0%", "¥595", "✅ 最高CPA効率"),
        ("no90-2", "画像", 1, "¥1,015", "3.56%", "6.67%", "¥1,015", "✅"),
        ("no91-2(検証)", "画像", 1, "¥1,442", "2.15%", "5.0%", "¥1,442", "✅"),
        ("no44(INT)", "画像", 1, "¥1,436", "2.85%", "7.14%", "¥1,436", "✅"),
        ("no98", "画像", 1, "¥2,559", "1.33%", "2.44%", "¥2,559", "✅"),
        ("no91-1", "画像", 1, "¥2,598", "3.47%", "4.0%", "¥2,598", "✅"),
        ("no91-2(ASC)", "画像", 1, "¥4,039", "2.84%", "3.03%", "¥4,039", "✅"),
        ("no56", "画像", 2, "¥4,737", "1.36%", "1.41%", "¥9,474", "✅"),
    ]
    for i, row_data in enumerate(axis_b_data):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        row += 1

    # Subtotal B
    subtotal_b = ("軸B合計（CV獲得分）", "-", 34, "¥2,655（平均）", "3.26%", "4.27%", "¥160,615", "-")
    for j, val in enumerate(subtotal_b, 1):
        cell = ws.cell(row=row, column=j, value=val)
        apply_data_style(cell, bg_color=LIGHT_GRAY, bold=True, size=9)
    row += 2

    # ---- Section C ----
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row=row, column=1, value="■ 訴求軸C：価格・お得感型（高CTR・低CVR問題群）")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    headers_c = ["CR名", "フォーマット", "CV数", "CPA", "CTR", "CVR", "消化金額", "問題点"]
    for i, h in enumerate(headers_c, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_header_style(cell, bg_color=LIGHT_BLUE)
    row += 1

    axis_c_data = [
        ("no9(ASC)", "画像", 3, "¥12,652", "2.63%", "0.58%", "¥37,957", "CTR並みだがCVR低すぎ"),
        ("no48(ASC検証)", "画像", 2, "¥13,134", "2.35%", "0.49%", "¥26,267", "同上"),
        ("no54_b(ASC)", "画像", 1, "¥32,943", "4.50%", "0.20%", "¥32,943", "CTR高いがCVR最悪"),
        ("no54_a(INT)", "画像", 0, "-", "6.94%", "0%", "¥1,168", "CTR6.9%なのにCV0→LP離脱"),
        ("no86(INT)", "画像", 0, "-", "3.57%", "0%", "¥2,342", "CV0"),
        ("no87(INT)", "画像", 0, "-", "4.29%", "0%", "¥1,530", "CV0"),
    ]
    for i, row_data in enumerate(axis_c_data):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        row += 1

    # Subtotal C
    subtotal_c = ("軸C合計", "-", 6, "¥19,576（平均）", "4.05%", "0.22%", "¥102,207", "コスト対比CV効率が最悪")
    for j, val in enumerate(subtotal_c, 1):
        cell = ws.cell(row=row, column=j, value=val)
        cell.fill = make_fill(LIGHT_RED)
        cell.font = make_font(bold=True, size=9)
        cell.alignment = make_align(horizontal="left")
        cell.border = make_border()
    row += 2

    # ---- Section D: Summary comparison ----
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row=row, column=1, value="■ 訴求軸別サマリー比較")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    summary_headers = ["訴求軸", "総CV", "平均CPA", "平均CTR", "平均CVR", "総消化金額", "評価", ""]
    for i, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_header_style(cell, bg_color=DARK_BLUE, font_color=WHITE)
    row += 1

    summary_data = [
        ("軸A 生活提案型", 26, "¥5,245", "2.29%", "1.69%", "¥134,973", "✅ 安定・スケール可能"),
        ("軸B 商品・品質型", 34, "¥2,655", "3.26%", "4.27%", "¥160,615", "✅ 高効率・主力候補"),
        ("軸C 価格・お得感型", 6, "¥19,576", "4.05%", "0.22%", "¥102,207", "❌ 高コスト低効率"),
    ]
    summary_colors = [LIGHT_GREEN, LIGHT_GREEN, LIGHT_RED]
    for row_data, color in zip(summary_data, summary_colors):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            apply_data_style(cell, bg_color=color, size=9, bold=True)
        cell = ws.cell(row=row, column=8)
        cell.fill = make_fill(color)
        cell.border = make_border()
        row += 1

    row += 1
    # Note
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = "※ 訴求軸Cは「広告でのクリックは集めるが、LPでの購入意思がマッチしない」パターン。解約理由1位「価格が高い」との連鎖リスクもあり、LPとの一貫性改善前提で予算抑制を推奨。"
    cell.font = make_font(italic=True, size=9, color="666666")
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 30

    # Column widths
    col_widths = {1: 24, 2: 14, 3: 8, 4: 14, 5: 8, 6: 8, 7: 14, 8: 40}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A3"
    return ws


# ==================== SHEET 4: 今後のCRアクションプラン ====================

def create_sheet4(wb):
    ws = wb.create_sheet("今後のCRアクションプラン")

    # Title
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "今後のMeta広告CRアクションプラン（2026年4月〜）"
    cell.fill = make_fill(DARK_BLUE)
    cell.font = make_font(bold=True, color=WHITE, size=14)
    cell.alignment = make_align(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    cell = ws["A2"]
    cell.value = "データ期間：2026/01/01〜2026/03/15 実績に基づく推奨アクション"
    cell.fill = make_fill(LIGHT_GRAY)
    cell.font = make_font(italic=True, size=10)
    cell.alignment = make_align(horizontal="center")

    row = 4

    # ---- Section A: 即実行アクション ----
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="■ 即実行アクション（〜2週間以内）")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    headers_a = ["#", "アクション", "対象", "期待効果", "担当"]
    for i, h in enumerate(headers_a, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_header_style(cell, bg_color=DARK_ORANGE, font_color=WHITE)
    row += 1

    immediate_actions = [
        (1, "no54_b（ASC）の停止", "no54_b", "¥32,943→年換算¥400K以上の無駄排除。予算をno85-4/動画_08に移管", "運用担当"),
        (2, "no89-1（全キャンペーン）の停止", "no89-1×3", "CV0確定。配信リソースの解放", "運用担当"),
        (3, "no88-2, no85-1, no90-1の停止", "各1本", "同系列の勝ちCRがある場合、負けCRは即停止", "運用担当"),
        (4, "INTキャンペーン全体の予算見直し", "INTキャンペーン", "34CR中1CVでCPA¥25,740。予算¥25,740→ASCに移管するだけでCVが増える計算", "運用担当"),
        (5, "no62の予算大幅増額テスト", "no62", "CPA¥595で最高効率。現状IMP141は少なすぎ。×10〜×50規模でテスト", "運用担当"),
        (6, "no90-2, no88-1の予算増額テスト", "no90-2, no88-1", "CPA¥1,015・¥2,358と高効率。スケールアップでCV数増加", "運用担当"),
    ]
    for i, row_data in enumerate(immediate_actions):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # ---- Section B: 短期アクション ----
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="■ 短期アクション（1〜2ヶ月）")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    headers_b = ["#", "アクション", "詳細", "目標KPI", ""]
    for i, h in enumerate(headers_b, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_header_style(cell, bg_color=DARK_BLUE, font_color=WHITE)
    row += 1

    short_actions = [
        (1, "訴求軸A（生活提案型）動画の新規制作", "動画_08の成功パターンを踏襲した新動画CR制作。「朝食シーン」「家族でパンを楽しむシーン」などライフスタイル訴求。目安：2〜3本", "CPA¥6,000以下"),
        (2, "訴求軸B（商品品質型）カルーセルの新規制作", "2405_カルーセル（CPA¥2,035〜¥2,878）が高効率。6〜8種類のパンを並べた新カルーセルを2〜3パターン制作", "CPA¥3,000以下"),
        (3, "no85-4系の派生CR制作", "no85-4の勝ち要素（おそらくビジュアル・コピー）を踏襲した派生CR。-1,-3が失敗した理由と-4の差分を分析", "CPA¥5,000以下"),
        (4, "no62の訴求を大きいフォーマットで再現", "no62（CTR7.09% CVR10%）のメッセージ・ビジュアルをストーリーズ縦型など別フォーマットでも展開", "CPA¥2,000以下"),
    ]
    for i, row_data in enumerate(short_actions):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        # Fill col 5
        cell = ws.cell(row=row, column=5)
        cell.fill = make_fill(bg)
        cell.border = make_border()
        ws.row_dimensions[row].height = 36
        row += 1

    row += 1

    # ---- Section C: 中期アクション ----
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1, value="■ 中期アクション（2〜4ヶ月）")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    headers_c = ["#", "アクション", "詳細", "目標", ""]
    for i, h in enumerate(headers_c, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_header_style(cell, bg_color=DARK_GREEN, font_color=WHITE)
    row += 1

    mid_actions = [
        (1, "UTMパラメーター設定（utm_content={{ad.name}}）", "代理店に依頼済み→確認・実施。CR別のBQ LTV追跡が可能になる", "CR別LTV測定開始"),
        (2, "訴求3軸のLTV検証", "UTM設定後、6ヶ月データが溜まったらCR訴求別の継続率・LTVを測定。「価格訴求→解約率高」の仮説を定量検証", "訴求軸別LTV比較"),
        (3, "twoMonthsプラン誘導LP/CRの制作", "twoMonthsの継続率94.7%（oneMonthより19pt高）。CRでtwoMonthsを前面に打ち出すか、LP上でtwoMonthsを選びやすくする", "twoMonths比率 15%以上"),
        (4, "季節・テーマ別CRの計画", "GW、夏、クリスマスなどの季節テーマに合わせたCRを事前制作。アドホック対応を脱却", "主要季節ごとに2〜3本"),
    ]
    for i, row_data in enumerate(mid_actions):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        cell = ws.cell(row=row, column=5)
        cell.fill = make_fill(bg)
        cell.border = make_border()
        ws.row_dimensions[row].height = 36
        row += 1

    row += 1

    # ---- Section D: 停止・予算削減リスト ----
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws.cell(row=row, column=1, value="■ 停止・予算削減リスト")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    headers_d = ["CR名", "キャンペーン", "現状CPA", "CV数", "消化金額", "停止理由"]
    for i, h in enumerate(headers_d, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_header_style(cell, bg_color=DARK_RED, font_color=WHITE)
    row += 1

    stop_data = [
        ("no54_b", "ASC", "¥32,943", 1, "¥32,943", "CPA許容値の5倍超。CTR高いがLP後離脱→LP改善が先"),
        ("no9", "ASC", "¥12,652", 3, "¥37,957", "CVR0.58%は低すぎ。予算をno85-4/動画_08に集中"),
        ("no48", "ASC検証", "¥13,134", 2, "¥26,267", "CPA高い。同予算で他CRの方がCV獲得効率高い"),
        ("no89-1", "ASC/ASC検証/INT", "CV0", 0, "¥2,988合計", "3キャンペーン全てで失敗。訴求不一致"),
        ("no88-2", "ASC/ASC検証", "CV0", 0, "¥4,457合計", "no88-1が機能→-2は不要"),
        ("no85-1", "ASC", "CV0", 0, "¥2,061", "no85-4が圧倒的に機能→-1は不要"),
        ("no90-1", "ASC/ASC検証", "CV0", 0, "¥2,797合計", "no90-2が機能→-1は不要"),
        ("INTキャンペーン大半", "INT", "¥25,740(全体)", 1, "¥25,740", "34CR中33CRがCV0。ASCへ予算移管"),
    ]
    for i, row_data in enumerate(stop_data):
        bg = LIGHT_RED
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            apply_data_style(cell, bg_color=bg, size=9)
        ws.row_dimensions[row].height = 26
        row += 1

    # Total stop budget
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = "合計停止予算: ¥139,213"
    cell.fill = make_fill(DARK_RED)
    cell.font = make_font(bold=True, color=WHITE, size=10)
    cell.alignment = make_align(horizontal="left")

    cell = ws.cell(row=row, column=6)
    cell.value = "この予算をASC主力CRに集中すれば理論上+22CV以上の追加獲得が可能"
    cell.fill = make_fill(DARK_RED)
    cell.font = make_font(bold=True, color=WHITE, size=9)
    cell.alignment = make_align(horizontal="left")
    cell.border = make_border()

    # Column widths
    col_widths = {1: 18, 2: 22, 3: 14, 4: 40, 5: 14, 6: 0}
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    ws.freeze_panes = "A3"
    return ws


# ==================== SHEET 5: CRネクスト制作ブリーフ ====================

def create_sheet5(wb):
    ws = wb.create_sheet("CRネクスト制作ブリーフ")

    # Title
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "新規CR制作ブリーフ（代理店提出用）"
    cell.fill = make_fill(DARK_BLUE)
    cell.font = make_font(bold=True, color=WHITE, size=14)
    cell.alignment = make_align(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    cell = ws["A2"]
    cell.value = "2026年4月以降の新規CR制作に向けた制作ブリーフ"
    cell.fill = make_fill(LIGHT_GRAY)
    cell.font = make_font(italic=True, size=10)
    cell.alignment = make_align(horizontal="center")

    row = 4

    def add_brief_section(ws, row, title, bg_color, data):
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
        cell.alignment = make_align(horizontal="left")
        ws.row_dimensions[row].height = 18
        row += 1

        # Headers
        for i, h in enumerate(["項目", "内容"], 1):
            cell = ws.cell(row=row, column=i, value=h)
            apply_header_style(cell, bg_color=LIGHT_BLUE)
        row += 1

        for i, (item, content) in enumerate(data):
            bg = WHITE if i % 2 == 0 else LIGHT_GRAY
            cell = ws.cell(row=row, column=1, value=item)
            apply_data_style(cell, bg_color=bg, size=9, bold=True)

            cell = ws.cell(row=row, column=2, value=content)
            apply_data_style(cell, bg_color=bg, size=9)
            ws.row_dimensions[row].height = 30
            row += 1

        return row + 1

    # Section A
    section_a_data = [
        ("優先度", "★★★（最高）"),
        ("フォーマット", "動画（縦型1080×1920 / 横型1080×1080）"),
        ("本数", "2〜3本"),
        ("参考CR", "動画_08（23CV CPA¥5,210）"),
        ("訴求コンセプト", "「冷凍庫にあるだけで、朝が変わる」「注文して届くまでのワクワク」「家族と食べる特別じゃない特別な朝食」"),
        ("必須要素", "①パンの焼き上がり・解凍シーン ②実際の食卓シーン（家族/一人/ペア） ③「毎月届く」「全国の職人から」などサービス説明"),
        ("避けるべき要素", "価格の数字を前面に出す（→軸Cとなりターゲット品質が下がる）"),
        ("KPI目標", "CPA ¥6,000以下（動画_08と同水準）"),
    ]
    row = add_brief_section(ws, row, "■ 最優先制作：訴求軸A 生活提案型 動画CR", DARK_ORANGE, section_a_data)

    # Section B
    section_b_data = [
        ("優先度", "★★★（最高）"),
        ("フォーマット", "カルーセル（1080×1080 × 5〜8枚）"),
        ("本数", "2〜3パターン"),
        ("参考CR", "2405_カルーセル01（CPA¥2,878）、2405_カルーセル02（CPA¥2,035）"),
        ("訴求コンセプト", "「今月届くパンはこれだ」品種のバラエティ・ビジュアルの訴求"),
        ("カルーセル構成案", "1枚目:ティザー「今月の顔ぶれ」→2〜7枚:各パンのアップ+産地/職人名→最終:CTA「今すぐ申し込む」"),
        ("KPI目標", "CPA ¥3,000以下"),
    ]
    row = add_brief_section(ws, row, "■ 優先制作：訴求軸B 商品・品質型 カルーセルCR", DARK_BLUE, section_b_data)

    # Section C
    section_c_data = [
        ("優先度", "★★（高）"),
        ("フォーマット", "静止画（1080×1080 / 1080×1350）"),
        ("本数", "3〜5本"),
        ("参考CR", "no62（CPA¥595 CVR10%）、no88-1（CPA¥2,358 CVR3.9%）、no90-2（CPA¥1,015 CVR6.7%）"),
        ("分析", "no62はCTR7.09%・CVR10%という異常値。このCRの「刺さっている要素」を代理店に確認して踏襲すること"),
        ("訴求コンセプト", "no62の要素を分析して記載（代理店と確認後）"),
        ("KPI目標", "CPA ¥3,000以下"),
    ]
    row = add_brief_section(ws, row, "■ 優先制作：訴求軸B 商品・品質型 静止画CR（no62/no88-1 系）", DARK_GREEN, section_c_data)

    # Section D: Stop list
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws.cell(row=row, column=1, value="■ 確認・廃止：停止推奨CR")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    stop_list = [
        "no54_b（ASC）: 消化¥32,943でCPA¥32,943。即停止",
        "no9（ASC）: 消化¥37,957でCPA¥12,652。予算削減",
        "no89-1（全キャンペーン）: 3配信全てCV0",
        "no88-2（ASC/ASC検証）: no88-1が機能するため不要",
        "INTキャンペーン: 34CR中1CV。抜本的な見直しまで予算停止",
    ]

    ws.merge_cells(f"A{row}:B{row}")
    cell = ws.cell(row=row, column=1, value="以下のCRは停止を代理店に依頼してください：")
    cell.fill = make_fill(LIGHT_RED)
    cell.font = make_font(bold=True, size=9)
    cell.alignment = make_align(horizontal="left")
    cell.border = make_border()
    ws.row_dimensions[row].height = 18
    row += 1

    for i, item in enumerate(stop_list):
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value=f"• {item}")
        bg = WHITE if i % 2 == 0 else LIGHT_RED
        cell.fill = make_fill(bg)
        cell.font = make_font(size=9)
        cell.alignment = make_align(horizontal="left")
        cell.border = make_border()
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # Section E: Confirmation items
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws.cell(row=row, column=1, value="■ 代理店への確認事項")
    cell.font = make_font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = make_align(horizontal="left")
    ws.row_dimensions[row].height = 18
    row += 1

    confirm_items = [
        ("1.", "no62のクリエイティブ内容確認：CTR7.09% CVR10%という驚異的な数値の要因を分析・共有してほしい"),
        ("2.", "no85-4のバリアント：-4が勝ちで-1/-3が負けの理由を分析・共有してほしい"),
        ("3.", "動画_08の成功要素：何が他の動画と違うのか整理してほしい"),
        ("4.", "UTMパラメーター設定：全Meta広告のURLに utm_content={{ad.name}} を追加設定してほしい（CR別LTV追跡のため）"),
        ("5.", "INTキャンペーンの今後方針：現状CPA¥25,740は許容不可。ターゲット設定の抜本見直しを提案してほしい"),
    ]

    for i, (num, content) in enumerate(confirm_items):
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY
        cell = ws.cell(row=row, column=1, value=num)
        apply_data_style(cell, bg_color=bg, size=9, bold=True)
        cell = ws.cell(row=row, column=2, value=content)
        apply_data_style(cell, bg_color=bg, size=9)
        ws.row_dimensions[row].height = 30
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    ws.freeze_panes = "A3"
    return ws


# ==================== MAIN ====================

def main():
    wb = openpyxl.Workbook()

    print("Creating Sheet 1: 訴求軸サマリー...")
    create_sheet1(wb)

    print("Creating Sheet 2: CR評価・アクション表...")
    create_sheet2(wb)

    print("Creating Sheet 3: 訴求別深掘り分析...")
    create_sheet3(wb)

    print("Creating Sheet 4: 今後のCRアクションプラン...")
    create_sheet4(wb)

    print("Creating Sheet 5: CRネクスト制作ブリーフ...")
    create_sheet5(wb)

    output_path = "/Users/mikiyakaneko/パンスク広告分析/パンスクCR訴求分析_アクションプラン.xlsx"
    wb.save(output_path)
    print(f"Saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    main()
