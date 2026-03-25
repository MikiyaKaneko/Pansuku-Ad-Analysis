#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

OUTPUT_PATH = "/Users/mikiyakaneko/パンスク広告分析/パンスク解約分析_レポート.xlsx"

# ── Color constants ──────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_LIGHT_BLUE  = "BDD7EE"
C_LIGHT_GRAY  = "F2F2F2"
C_LIGHT_YELLOW= "FFF2CC"
C_RED         = "FFE0E0"
C_ORANGE      = "FFF2CC"
C_GREEN       = "E2EFDA"
C_WHITE       = "FFFFFF"
C_HEADER_GRAY = "D9D9D9"

# ── Helper builders ──────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)

def apply_title_row(ws, row, text, merged_cols="A:H", bg=C_DARK_BLUE, txt=C_WHITE, size=14):
    # merged_cols should be like "A:H" (column letters only, no row numbers)
    parts = merged_cols.split(":")
    col_start = parts[0].strip("0123456789")
    col_end   = parts[1].strip("0123456789")
    ws.merge_cells(f"{col_start}{row}:{col_end}{row}")
    cell = ws[f"{col_start}{row}"]
    cell.value = text
    cell.font  = font(bold=True, color=txt, size=size)
    cell.fill  = fill(bg)
    cell.alignment = center_align(wrap=True)

def apply_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font      = font(bold=True, color="000000", size=11)
        c.fill      = fill(C_LIGHT_BLUE)
        c.alignment = center_align(wrap=True)
        c.border    = thin_border()

def apply_data_row(ws, row, values, start_col=1, bg=None, pct_cols=None):
    """Write a data row. pct_cols: set of 0-based column indices that are percentages."""
    pct_cols = pct_cols or set()
    row_bg = bg if bg else (C_WHITE if row % 2 == 0 else C_LIGHT_GRAY)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font      = font()
        c.fill      = fill(row_bg)
        c.alignment = left_align(wrap=True)
        c.border    = thin_border()
        if i in pct_cols and isinstance(v, (int, float)):
            c.number_format = "0.0%"

def set_col_widths(ws, widths):
    """widths: list of (col_letter_or_index, width)"""
    for col, w in widths:
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ════════════════════════════════════════════════════════════════════════════
# Sheet 1: サマリー
# ════════════════════════════════════════════════════════════════════════════

def build_summary(ws):
    ws.title = "サマリー"

    # Row 1: Title
    apply_title_row(ws, 1, "パンスク 解約分析レポート", "A1:H1", size=16)
    set_row_height(ws, 1, 30)

    # Row 2: Data period subtitle
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "データ期間：2025/03〜2026/03 / 対象：registerCompleted=true（決済完了者）"
    c.font  = font(italic=True, size=10)
    c.fill  = fill(C_LIGHT_GRAY)
    c.alignment = center_align()

    # Row 4: KPI section header
    c = ws["A4"]
    c.value = "■ 主要KPI"
    c.font  = font(bold=True, color=C_DARK_BLUE, size=12)

    # Row 5: KPI table header
    kpi_headers = ["指標", "値", "比較対象", "評価"]
    apply_header_row(ws, 5, kpi_headers, start_col=1)
    # extend fill to H
    for col in range(5, 9):
        c = ws.cell(row=5, column=col)
        c.fill   = fill(C_LIGHT_BLUE)
        c.border = thin_border()

    # Row 6-10: KPI data
    kpi_data = [
        ("Meta ASC 解約率",            "21.4%", "Google organic 23.8%",        "✅ 良好"),
        ("panforyou.jp 解約率",         "12.5%", "全チャネル最良",                "✅ 優秀"),
        ("aeoncard_202310 解約率",      "85.7%", "alliance-coupon最悪",          "❌ 即停止検討"),
        ("Meta ASC 解約まで平均日数",   "85日",  "kepco.jp 48日（早期離脱多）",  "⚠️ 要注意"),
        ("クーポン即解約（平均日数）",  "2〜18日","入会直後解約パターン",          "❌ LTV毀損"),
    ]
    for i, row_data in enumerate(kpi_data):
        r = 6 + i
        row_bg = C_WHITE if i % 2 == 0 else C_LIGHT_GRAY
        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font      = font()
            c.fill      = fill(row_bg)
            c.alignment = left_align(wrap=True)
            c.border    = thin_border()

    # Row 12: 解約理由 section header
    c = ws["A12"]
    c.value = "■ 解約理由TOP5カテゴリ"
    c.font  = font(bold=True, color=C_DARK_BLUE, size=12)

    # Row 13: column headers
    reason_headers = ["カテゴリ", "解約理由例", "特徴"]
    apply_header_row(ws, 13, reason_headers, start_col=1)

    # Row 14-18: data
    reason_data = [
        ("価格・コスト系",      "高い/値段が高い/料金が高い/経済的理由",             "解約理由の推定50%超・断トツ1位"),
        ("クーポン使えない",    "クーポンが使えなかった/クーポン併用不可",           "入会後2〜18日で即解約。LTV毀損"),
        ("必要なくなった",      "必要なくなった/利用しなくなった/不要",             "ライフスタイル変化"),
        ("品質・嗜好",          "好みのパンではなかった/美味しくない/飽きた",       "商品改善で対処余地あり"),
        ("食べきれない系",      "食べきれない/冷凍庫に入らない",                   "サービス特性起因。オンボーディングで対処可"),
    ]
    for i, row_data in enumerate(reason_data):
        r = 14 + i
        row_bg = C_WHITE if i % 2 == 0 else C_LIGHT_GRAY
        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font      = font()
            c.fill      = fill(row_bg)
            c.alignment = left_align(wrap=True)
            c.border    = thin_border()

    # Row 21: action section header
    c = ws["A21"]
    c.value = "■ アクション提言"
    c.font  = font(bold=True, color=C_DARK_BLUE, size=12)

    # Row 22-27: action items (merged A:H per row)
    actions = [
        "① 価格訴求の見直し: 解約理由1位が「価格が高い」→ 広告・LPでの価値訴求強化（値下げでなく価値の伝え方）",
        "② alliance-couponキャンペーン精査: aeoncard解約率85.7%は即停止。goopan500(14%)/persona_HP(0%)は継続",
        "③ kepco.jp連携条件の確認: 月別解約率が0%〜67%と不安定。2025-06/10に解約率50%超の原因調査",
        "④ オンボーディング改善: 「食べきれない」「冷凍庫に入らない」への対処（初回量説明・保存方法案内）",
        "⑤ twoMonthsプラン誘導強化: 継続率94.7%でoneMonthより19pt高い。LP訴求の見直し",
        "⑥ Meta ASCクリエイティブ: 2025-08が解約率14.3%と最良。その月のCRを参考に新CR開発",
    ]
    for i, action in enumerate(actions):
        r = 22 + i
        ws.merge_cells(f"A{r}:H{r}")
        c = ws[f"A{r}"]
        c.value = action
        c.font  = font(size=11)
        c.fill  = fill("FFFDE7")  # light yellow
        c.alignment = left_align(wrap=True)
        c.border = thin_border()
        set_row_height(ws, r, 22)

    # Column widths
    set_col_widths(ws, [
        (1, 28), (2, 18), (3, 30), (4, 15),
        (5, 10), (6, 10), (7, 10), (8, 10),
    ])

    ws.freeze_panes = "A3"


# ════════════════════════════════════════════════════════════════════════════
# Sheet 2: 解約理由分析
# ════════════════════════════════════════════════════════════════════════════

def build_leave_reason(ws):
    ws.title = "解約理由分析"

    # Title
    apply_title_row(ws, 1, "解約理由（leaveReason）分析", "A1:D1", size=14)
    set_row_height(ws, 1, 28)

    # Headers
    headers = ["解約理由カテゴリ", "代表的な表記", "件数（推定）", "特徴・考察"]
    apply_header_row(ws, 2, headers, start_col=1)

    # Category data: (category, reason_text, count_text, note)
    categories = [
        # 価格・コスト系
        ("価格・コスト系", "高い",                 "207件",  ""),
        ("",              "値段が高い",             "143件",  ""),
        ("",              "料金が高い",              "69件",   ""),
        ("",              "高いから",               "67件",   ""),
        ("",              "価格が高い",              "50件",   ""),
        ("",              "値段が高いため",           "49件",   ""),
        ("",              "金額が高い",              "48件",   ""),
        ("",              "経済的理由",             "30件",   ""),
        ("",              "※その他「値段が高いから」「料金が高いため」等多数の表記ゆれあり", "-件", ""),
        # subtotal
        ("【合計】価格・コスト系", "価格・コスト系 合計（推定）", "約1,200件以上", "解約理由の推定50%超。値下げよりも価値訴求の見直しが有効"),

        # 無記入
        ("無記入・形式的入力", "NULL（未記入）",  "2,905件", ""),
        ("",                  "空文字",           "2,584件", ""),
        ("",                  "※「あ」「。」「、」等の形式的入力も含む", "-件", ""),
        ("【合計】無記入・形式的入力", "合計",     "約5,500件", "全解約者の約22%。フォーム設計の改善余地あり"),

        # クーポン
        ("クーポン使えない系", "クーポンが使えなかったため", "26件", ""),
        ("",                   "クーポン併用不可のため",       "19件", ""),
        ("",                   "クーポンが使えなかったから",   "17件", ""),
        ("",                   "クーポンが利用できなかったため","13件", ""),
        ("",                   "クーポンが使えないため",       "13件", ""),
        ("",                   "※その他類似表記多数",          "-件",  ""),
        ("【合計】クーポン使えない系", "合計", "約260件", "解約まで平均2〜18日。入会直後に解約するパターン。LTVを大幅に毀損"),

        # 必要なくなった
        ("必要なくなった系", "必要なくなったため",   "14件", ""),
        ("",                 "利用しなくなったため", "15件", ""),
        ("",                 "必要なくなった",       "14件", ""),
        ("",                 "不要",                 "9件",  ""),
        ("【合計】必要なくなった系", "合計", "約110件", "ライフスタイル変化による自然離脱"),

        # 品質・嗜好
        ("品質・嗜好系", "好みのパンが少ない",     "14件", ""),
        ("",             "美味しくなかった",        "13件", ""),
        ("",             "好みのパンではなかった",  "10件", ""),
        ("",             "あまり美味しくない",       "9件", ""),
        ("",             "飽きた",                   "8件", ""),
        ("【合計】品質・嗜好系", "合計", "約90件", "品種ラインアップ・品質改善で対処余地あり"),

        # 食べきれない
        ("食べきれない系", "食べきれない",     "20件", ""),
        ("",               "食べきれないため", "13件", ""),
        ("",               "冷凍庫に入らない",  "9件", ""),
        ("【合計】食べきれない系", "合計", "約60件", "冷凍ボックス特有の課題。初回量説明・プラン案内で軽減可能"),

        # ライフスタイル変化
        ("ライフスタイル変化", "ダイエットのため", "13件", ""),
        ("",                   "体調不良の為",      "9件", ""),
        ("",                   "引越しのため",      "8件", ""),
        ("【合計】ライフスタイル変化", "合計", "約30件", "外部要因による離脱（対処困難）"),

        # 満足・卒業
        ("満足・卒業系", "満足したため",          "12件", ""),
        ("",             "満足したから",          "12件", ""),
        ("",             "ありがとうございました","12件", ""),
        ("【合計】満足・卒業系", "合計", "約50件", "ポジティブな離脱（サービスへの感謝あり）"),
    ]

    # Define category color bands
    cat_colors = {
        "価格・コスト系":    "FFF9C4",
        "無記入・形式的入力": "E8EAF6",
        "クーポン使えない系": "FCE4EC",
        "必要なくなった系":  "E8F5E9",
        "品質・嗜好系":      "FFF3E0",
        "食べきれない系":    "E1F5FE",
        "ライフスタイル変化": "F3E5F5",
        "満足・卒業系":      "E0F2F1",
    }
    subtotal_color = "EEEEEE"

    current_cat = None
    current_color = C_WHITE

    for i, (cat, reason, count, note) in enumerate(categories):
        r = 3 + i
        is_subtotal = cat.startswith("【合計】")

        # Determine category
        if cat and not is_subtotal:
            current_cat = cat
            for key, color in cat_colors.items():
                if key in cat:
                    current_color = color
                    break

        display_cat = cat if not is_subtotal else ""
        if is_subtotal:
            row_bg = subtotal_color
        else:
            row_bg = current_color

        for j, v in enumerate([display_cat if not is_subtotal else cat.replace("【合計】","→ "), reason, count, note]):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font      = font(bold=is_subtotal, size=10)
            c.fill      = fill(row_bg)
            c.alignment = left_align(wrap=True)
            c.border    = thin_border()

    # Note at end
    last_row = 3 + len(categories) + 1
    ws.merge_cells(f"A{last_row}:D{last_row}")
    c = ws[f"A{last_row}"]
    c.value = "※ leaveReasonはフリーテキスト入力のため、同一意味の表記ゆれが多数存在。件数は代表的な表記のみの数値であり実際の総数はより多い。"
    c.font  = font(italic=True, size=9, color="555555")
    c.alignment = left_align(wrap=True)
    set_row_height(ws, last_row, 30)

    set_col_widths(ws, [(1, 25), (2, 40), (3, 15), (4, 45)])
    ws.freeze_panes = "A3"


# ════════════════════════════════════════════════════════════════════════════
# Sheet 3: 流入源別解約率
# ════════════════════════════════════════════════════════════════════════════

def build_source_churn(ws):
    ws.title = "流入源別解約率"

    # Title
    apply_title_row(ws, 1, "流入源別 解約率分析", "A1:I1", size=14)
    set_row_height(ws, 1, 28)

    # Section A header
    headers_a = ["流入源", "キャンペーン", "総ユーザー数", "解約数", "休止中", "アクティブ", "解約率", "平均継続月数", "評価"]
    apply_header_row(ws, 2, headers_a, start_col=1)

    # Data sorted by churn_rate desc
    source_data = [
        # source, campaign, total, churned, paused, active, churn_rate, avg_months, eval
        ("alliance-coupon", "aeoncard_202310",   7,   6, 0,  1,  0.857, "5.1ヶ月",  "❌ 即停止検討"),
        ("alliance-coupon", "kanden_202312",     16,  8, 0,  8,  0.500, "10.9ヶ月", "❌ 要見直し"),
        ("alliance-coupon", "ruum_202304",        7,  3, 0,  4,  0.429, "7.6ヶ月",  "⚠️"),
        ("alliance-coupon", "viewpan1000",        5,  2, 0,  3,  0.400, "6.2ヶ月",  "⚠️"),
        ("aumo.jp",         "(referral)",         5,  2, 0,  3,  0.400, "7.2ヶ月",  "⚠️"),
        ("kepco.jp",        "(referral)",        73, 25, 0, 48,  0.342, "5.7ヶ月",  "⚠️ 月別ブレ大"),
        ("yahoo",           "(organic)",         35,  9, 0, 26,  0.257, "7.5ヶ月",  "普通"),
        ("ads",             "newspaper_5d_2024", 21,  6, 0, 15,  0.286, "11.4ヶ月", "普通"),
        ("alliance-coupon", "ucs_202307",         7,  2, 0,  5,  0.286, "5.4ヶ月",  "普通"),
        ("alliance-coupon", "ponta_2404",         7,  2, 0,  5,  0.286, "5.9ヶ月",  "普通"),
        ("alliance-coupon", "jfrcard_202306",     9,  3, 0,  6,  0.333, "6.1ヶ月",  "⚠️"),
        ("google",          "(organic)",        143, 34, 1,108,  0.238, "7.7ヶ月",  "普通"),
        ("alliance-coupon", "persona_LINE",      30,  7, 0, 23,  0.233, "7.0ヶ月",  "普通"),
        ("alliancecoupon",  "jqcard_2503",        4,  1, 0,  3,  0.250, "12.0ヶ月", "普通"),
        ("alliance-coupon", "bluerosecard25",     4,  1, 0,  3,  0.250, "11.8ヶ月", "普通"),
        ("meta",            "ASC",             168, 36, 0,132,  0.214, "5.8ヶ月",  "✅ 良好"),
        ("bing",            "(organic)",         5,  1, 0,  4,  0.200, "6.6ヶ月",  "✅"),
        ("meta",            "INT",              20,  3, 0, 17,  0.150, "9.2ヶ月",  "✅"),
        ("ktv.jp",          "(referral)",        6,  1, 0,  5,  0.167, "4.0ヶ月",  "✅"),
        ("kinami-bread.com","(referral)",       67, 11, 0, 56,  0.164, "5.7ヶ月",  "✅ 良好"),
        ("(direct)",        "(direct)",         58,  9, 0, 49,  0.155, "7.0ヶ月",  "✅ 良好"),
        ("alliance-coupon", "goopan500",         7,  1, 0,  6,  0.143, "3.7ヶ月",  "✅"),
        ("panforyou.jp",    "(referral)",       24,  3, 0, 21,  0.125, "8.9ヶ月",  "✅ 最良"),
        ("smartnews.com",   "(referral)",        4,  0, 0,  4,  0.000, "11.0ヶ月", "✅"),
        ("alliance-coupon", "persona_HP",        3,  0, 0,  3,  0.000, "7.3ヶ月",  "✅"),
    ]

    # Sort by churn rate desc
    source_data_sorted = sorted(source_data, key=lambda x: x[6], reverse=True)

    for i, row_data in enumerate(source_data_sorted):
        r = 3 + i
        churn_rate = row_data[6]

        if churn_rate > 0.50:
            row_bg = "FFE0E0"
        elif churn_rate >= 0.30:
            row_bg = "FFF2CC"
        elif churn_rate < 0.20:
            row_bg = "E2EFDA"
        else:
            row_bg = C_WHITE if i % 2 == 0 else C_LIGHT_GRAY

        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font      = font(size=10)
            c.fill      = fill(row_bg)
            c.alignment = center_align() if j >= 2 else left_align()
            c.border    = thin_border()
            if j == 6:  # churn_rate column
                c.number_format = "0.0%"

    # Section B: avg days to churn
    b_start = 3 + len(source_data_sorted) + 2
    ws.cell(row=b_start - 1, column=1).value = "解約までの平均日数（流入源別）"
    ws.cell(row=b_start - 1, column=1).font  = font(bold=True, color=C_DARK_BLUE, size=12)

    headers_b = ["流入源", "解約件数", "解約まで平均日数", "解釈"]
    apply_header_row(ws, b_start, headers_b, start_col=1)

    days_data = [
        ("kepco.jp",          24, "48日",  "⚠️ 早期離脱が多い。初回クーポン目的の可能性"),
        ("yahoo",              9, "62日",  "やや短め"),
        ("meta ASC",          30, "85日",  "約2.8ヶ月で解約決着"),
        ("google organic",    29, "83日",  "meta ASCと同水準"),
        ("(direct)",           8, "93日",  "比較的長く使ってから解約"),
        ("kinami-bread.com",  10, "94日",  "最も長く使ってから解約"),
    ]
    for i, row_data in enumerate(days_data):
        r = b_start + 1 + i
        row_bg = C_WHITE if i % 2 == 0 else C_LIGHT_GRAY
        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font      = font(size=10)
            c.fill      = fill(row_bg)
            c.alignment = left_align(wrap=True)
            c.border    = thin_border()

    set_col_widths(ws, [
        (1, 20), (2, 25), (3, 12), (4, 10), (5, 10),
        (6, 12), (7, 10), (8, 14), (9, 20)
    ])
    ws.freeze_panes = "A3"


# ════════════════════════════════════════════════════════════════════════════
# Sheet 4: 月別コホート
# ════════════════════════════════════════════════════════════════════════════

def build_cohort(ws):
    ws.title = "月別コホート"

    # Title
    apply_title_row(ws, 1, "月別解約コホート分析（2025/03〜2026/03）", "A1:E1", size=14)
    set_row_height(ws, 1, 28)

    # Subtitle
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = "※2025年3月以降はGA4追跡可能。それ以前はほぼ unknown"
    c.font  = font(italic=True, size=10, color="555555")
    c.fill  = fill(C_LIGHT_GRAY)
    c.alignment = left_align()

    # ── Section A: Meta ASC ─────────────────────────────────────────────────
    ws.cell(row=4, column=1).value = "▼ Meta ASC 月別解約率推移"
    ws.cell(row=4, column=1).font  = font(bold=True, color=C_DARK_BLUE, size=12)

    headers = ["登録月", "総数", "解約数", "解約率", "評価コメント"]
    apply_header_row(ws, 5, headers, start_col=1)

    meta_data = [
        ("2025-03", 27,  5, 0.185, "基準月"),
        ("2025-04",  7,  3, 0.429, "⚠️ 高め"),
        ("2025-05", 11,  4, 0.364, "⚠️ 高め"),
        ("2025-06",  9,  5, 0.556, "❌ 高い（kepco.jpも同月高）"),
        ("2025-07",  6,  2, 0.333, "⚠️"),
        ("2025-08", 14,  2, 0.143, "✅ 最良月"),
        ("2025-09", 13,  4, 0.308, "⚠️"),
        ("2025-10", 10,  4, 0.400, "❌ 高い（kepco.jpも同月高）"),
        ("2025-11",  8,  2, 0.250, "普通"),
        ("2025-12", 13,  3, 0.231, "普通"),
        ("2026-01", 15,  1, 0.067, "✅ まだ新しい"),
        ("2026-02", 16,  3, 0.188, "✅ まだ新しい"),
        ("2026-03", 14,  0, 0.000, "当月登録"),
    ]

    for i, row_data in enumerate(meta_data):
        r = 6 + i
        churn_rate = row_data[3]
        if churn_rate > 0.50:
            row_bg = "FFE0E0"
        elif churn_rate >= 0.30:
            row_bg = "FFF2CC"
        elif churn_rate < 0.20:
            row_bg = "E2EFDA"
        else:
            row_bg = C_WHITE if i % 2 == 0 else C_LIGHT_GRAY

        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font      = font(size=10)
            c.fill      = fill(row_bg)
            c.alignment = center_align() if j in (1, 2, 3) else left_align()
            c.border    = thin_border()
            if j == 3:
                c.number_format = "0.0%"

    # ── Section B: kepco.jp ─────────────────────────────────────────────────
    b_row = 6 + len(meta_data) + 2
    ws.cell(row=b_row - 1, column=1).value = "▼ kepco.jp 月別解約率推移（ブレ確認）"
    ws.cell(row=b_row - 1, column=1).font  = font(bold=True, color=C_DARK_BLUE, size=12)

    apply_header_row(ws, b_row, headers, start_col=1)

    kepco_data = [
        ("2025-06", 19, 9, 0.474, "❌ 高い"),
        ("2025-07",  3, 2, 0.667, "❌ 少数だが高い"),
        ("2025-09", 13, 4, 0.308, "⚠️"),
        ("2025-10", 11, 6, 0.545, "❌ 高い"),
        ("2025-11",  4, 1, 0.250, "普通"),
        ("2025-12",  5, 1, 0.200, "普通"),
        ("2026-01",  4, 0, 0.000, "まだ新しい"),
        ("2026-02",  4, 0, 0.000, "まだ新しい"),
        ("2026-03",  4, 1, 0.250, "まだ新しい"),
    ]

    for i, row_data in enumerate(kepco_data):
        r = b_row + 1 + i
        churn_rate = row_data[3]
        if churn_rate > 0.50:
            row_bg = "FFE0E0"
        elif churn_rate >= 0.30:
            row_bg = "FFF2CC"
        elif churn_rate < 0.20:
            row_bg = "E2EFDA"
        else:
            row_bg = C_WHITE if i % 2 == 0 else C_LIGHT_GRAY

        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font      = font(size=10)
            c.fill      = fill(row_bg)
            c.alignment = center_align() if j in (1, 2, 3) else left_align()
            c.border    = thin_border()
            if j == 3:
                c.number_format = "0.0%"

    # ── Section C: 2025-03コホート比較 ─────────────────────────────────────
    c_row = b_row + 1 + len(kepco_data) + 2
    ws.cell(row=c_row - 1, column=1).value = "▼ 流入源別 同月コホート比較（2025-03登録コホート）"
    ws.cell(row=c_row - 1, column=1).font  = font(bold=True, color=C_DARK_BLUE, size=12)

    headers_c = ["流入源", "総数", "解約数", "解約率", "評価"]
    apply_header_row(ws, c_row, headers_c, start_col=1)

    cohort_data = [
        ("meta",              27, 5, 0.185, "✅"),
        ("google",            20, 5, 0.250, "普通"),
        ("ads（新聞）",       14, 5, 0.357, "⚠️"),
        ("unknown",           13, 4, 0.308, "参考値"),
        ("alliance-coupon",   11, 5, 0.455, "❌"),
        ("(direct)",           6, 0, 0.000, "✅（少数）"),
        ("kinami-bread.com",   5, 1, 0.200, "✅"),
        ("panforyou.jp",       3, 0, 0.000, "✅（少数）"),
    ]

    for i, row_data in enumerate(cohort_data):
        r = c_row + 1 + i
        churn_rate = row_data[3]
        if churn_rate > 0.50:
            row_bg = "FFE0E0"
        elif churn_rate >= 0.30:
            row_bg = "FFF2CC"
        elif churn_rate < 0.20:
            row_bg = "E2EFDA"
        else:
            row_bg = C_WHITE if i % 2 == 0 else C_LIGHT_GRAY

        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font      = font(size=10)
            c.fill      = fill(row_bg)
            c.alignment = center_align() if j in (1, 2, 3) else left_align()
            c.border    = thin_border()
            if j == 3:
                c.number_format = "0.0%"

    # Note
    note_row = c_row + 1 + len(cohort_data) + 1
    ws.merge_cells(f"A{note_row}:E{note_row}")
    c = ws[f"A{note_row}"]
    c.value = "→ 2025-03コホートではMeta ASCの解約率18.5%は全チャネル中2番目に低い。alliance-coupon(45.5%)やads(35.7%)より明らかに優秀。"
    c.font  = font(italic=True, size=10, color="1F3864")
    c.fill  = fill("EBF3FB")
    c.alignment = left_align(wrap=True)
    set_row_height(ws, note_row, 28)

    set_col_widths(ws, [(1, 22), (2, 10), (3, 10), (4, 10), (5, 38)])
    ws.freeze_panes = "A3"


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws1 = wb.create_sheet("サマリー")
    ws2 = wb.create_sheet("解約理由分析")
    ws3 = wb.create_sheet("流入源別解約率")
    ws4 = wb.create_sheet("月別コホート")

    build_summary(ws1)
    build_leave_reason(ws2)
    build_source_churn(ws3)
    build_cohort(ws4)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
